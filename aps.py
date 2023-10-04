import tkinter as tk
from tkinter import messagebox
import openpyxl 


def adicionar_cliente():
    nome = entry_nome.get()
    email = entry_email.get()
    telefone = entry_telefone.get()
    
    if nome and email and telefone:
        global codigo_cliente
        codigo_cliente += 1
        codigo = codigo_cliente
        
     
        wb = openpyxl.load_workbook('clientes.xlsx')
        ws = wb.active
        ws.append([codigo, nome, email, telefone])
        wb.save('clientes.xlsx')
        
        messagebox.showinfo("Cadastro de Cliente", "Cliente cadastrado com sucesso!")
        
      
        entry_nome.delete(0, tk.END)
        entry_email.delete(0, tk.END)
        entry_telefone.delete(0, tk.END)
    else:
        messagebox.showwarning("Cadastro de Cliente", "Por favor, preencha todos os campos.")


def pesquisar_cliente():
    nome_pesquisa = entry_pesquisa.get()
    if nome_pesquisa:
        wb = openpyxl.load_workbook('clientes.xlsx')
        ws = wb.active
        
        for row in ws.iter_rows(values_only=True):
            codigo, nome, email, telefone = row
            if nome == nome_pesquisa:
                messagebox.showinfo("Informações do Cliente",
                                    f"Código: {codigo}\nNome: {nome}\nEmail: {email}\nTelefone: {telefone}")
                break
        else:
            messagebox.showwarning("Pesquisa de Cliente", f"Cliente com nome '{nome_pesquisa}' não encontrado.")
    else:
        messagebox.showwarning("Pesquisa de Cliente", "Por favor, insira um nome para pesquisa.")


def editar_cliente():
    codigo_edicao = entry_codigo_edicao.get()
    nome_edicao = entry_nome_edicao.get()
    email_edicao = entry_email_edicao.get()
    telefone_edicao = entry_telefone_edicao.get()
    
    if codigo_edicao and nome_edicao and email_edicao and telefone_edicao:
        wb = openpyxl.load_workbook('clientes.xlsx')
        ws = wb.active
        
        for row in ws.iter_rows(values_only=True):
            codigo, nome, email, telefone = row
            if codigo == int(codigo_edicao):
                row[1] = nome_edicao
                row[2] = email_edicao
                row[3] = telefone_edicao
                wb.save('clientes.xlsx')
                messagebox.showinfo("Edição de Cliente", "Cliente editado com sucesso!")
                break
        else:
            messagebox.showwarning("Edição de Cliente", f"Cliente com código {codigo_edicao} não encontrado.")
    else:
        messagebox.showwarning("Edição de Cliente", "Por favor, preencha todos os campos de edição.")

root = tk.Tk()
root.title("Cadastro de Clientes")

codigo_cliente = 0

label_nome = tk.Label(root, text="Nome:")
entry_nome = tk.Entry(root)

label_email = tk.Label(root, text="Email:")
entry_email = tk.Entry(root)

label_telefone = tk.Label(root, text="Telefone:")
entry_telefone = tk.Entry(root)

botao_adicionar = tk.Button(root, text="Adicionar Cliente", command=adicionar_cliente)

label_pesquisa = tk.Label(root, text="Pesquisar por Nome:")
entry_pesquisa = tk.Entry(root)
botao_pesquisar = tk.Button(root, text="Pesquisar Cliente", command=pesquisar_cliente)

label_codigo_edicao = tk.Label(root, text="Código do Cliente para Edição:")
entry_codigo_edicao = tk.Entry(root)

label_nome_edicao = tk.Label(root, text="Novo Nome:")
entry_nome_edicao = tk.Entry(root)

label_email_edicao = tk.Label(root, text="Novo Email:")
entry_email_edicao = tk.Entry(root)

label_telefone_edicao = tk.Label(root, text="Novo Telefone:")
entry_telefone_edicao = tk.Entry(root)

botao_editar = tk.Button(root, text="Editar Cliente", command=editar_cliente)

label_nome.grid(row=0, column=0)
entry_nome.grid(row=0, column=1)

label_email.grid(row=1, column=0)
entry_email.grid(row=1, column=1)

label_telefone.grid(row=2, column=0)
entry_telefone.grid(row=2, column=1)

botao_adicionar.grid(row=3, column=0, columnspan=2)

label_pesquisa.grid(row=4, column=0)
entry_pesquisa.grid(row=4, column=1)
botao_pesquisar.grid(row=4, column=2)

label_codigo_edicao.grid(row=5, column=0)
entry_codigo_edicao.grid(row=5, column=1)

label_nome_edicao.grid(row=6, column=0)
entry_nome_edicao.grid(row=6, column=1)

label_email_edicao.grid(row=7, column=0)
entry_email_edicao.grid(row=7, column=1)

label_telefone_edicao.grid(row=8, column=0)
entry_telefone_edicao.grid(row=8, column=1)

botao_editar.grid(row=9, column=0, columnspan=2)

root.mainloop()